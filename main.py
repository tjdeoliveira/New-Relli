from flask import Flask, request, render_template_string, Response
import requests
import csv
from datetime import datetime, timezone, timedelta
import io
import os

app = Flask(__name__)

API_KEY = os.getenv('NEW_RELIC_API_KEY')
ACCOUNT_ID = os.getenv('NEW_RELIC_ACCOUNT_ID')
MAX_RESULTS = 5000  # Limite máximo de resultados por consulta NRQL
MIN_INTERVAL = timedelta(minutes=1)  # Intervalo mínimo para evitar recursão infinita

# Validar variáveis de ambiente
if not API_KEY or not ACCOUNT_ID:
    raise EnvironmentError(
        "NEW_RELIC_API_KEY ou NEW_RELIC_ACCOUNT_ID não estão definidos. Verifique suas variáveis de ambiente.")


def parse_brazilian_datetime(date_str):
    """Parse a datetime string from datetime-local to UTC datetime."""
    try:
        dt_br = datetime.strptime(date_str, '%Y-%m-%dT%H:%M')
        tz_br = timezone(timedelta(hours=-3))
        dt_br = dt_br.replace(tzinfo=tz_br)
        dt_utc = dt_br.astimezone(timezone.utc)
        return dt_utc, None
    except ValueError as e:
        return None, f"Erro ao parsear data: {e}. Use o formato fornecido pelo campo de data."


def fetch_newrelic_data(start_time, end_time, company_id):
    """Fetch data from New Relic for a given time range and company ID."""
    nrql_query = (
        "SELECT * "
        "FROM Log "
        "WHERE (type = 'SENT_MESSAGE_STATUS' OR type = 'SENT_MESSAGE' OR type = 'SENT_ALL_MESSAGE') "
        f"AND (status.code IN ('failed') AND company.id IN ('{company_id}')) "
        f"SINCE '{start_time.strftime('%Y-%m-%d %H:%M:%S')} UTC' "
        f"UNTIL '{end_time.strftime('%Y-%m-%d %H:%M:%S')} UTC' "
        "LIMIT MAX "
        "ORDER BY timestamp ASC"
    )
    payload = {
        "query": f"""{{
          actor {{
            account(id: {ACCOUNT_ID}) {{
              nrql(query: "{nrql_query}") {{
                results
              }}
            }}
          }}
        }}"""
    }
    try:
        response = requests.post(
            "https://api.newrelic.com/graphql",
            headers={
                "API-Key": API_KEY,
                "Content-Type": "application/json"
            },
            json=payload
        )
        response.raise_for_status()  # Levanta exceção para códigos de erro HTTP
        try:
            data = response.json()
        except ValueError:
            return [], f"Erro: Resposta da API não é um JSON válido ({response.status_code})"

        # Verificar se a estrutura esperada existe
        if not isinstance(data, dict):
            return [], f"Erro: Resposta da API não contém um dicionário válido"

        results = data.get("data", {}).get("actor", {}).get("account", {}).get("nrql", {}).get("results", [])
        return results, None
    except requests.RequestException as e:
        return [], f"Erro na requisição: {str(e)}"


def fetch_recursive(start_time, end_time, company_id):
    """Recursive fetch with intelligent subdivision if limit is exceeded."""
    if end_time - start_time <= MIN_INTERVAL:
        return fetch_newrelic_data(start_time, end_time, company_id)

    results, error = fetch_newrelic_data(start_time, end_time, company_id)
    if error:
        return results, error
    if len(results) < MAX_RESULTS:
        return results, None
    else:
        mid_time = start_time + (end_time - start_time) / 2
        left_results, left_error = fetch_recursive(start_time, mid_time, company_id)
        if left_error:
            return left_results, left_error
        right_results, right_error = fetch_recursive(mid_time, end_time, company_id)
        if right_error:
            return right_results, right_error
        return left_results + right_results, None


HTML_MESSAGE_ID = """
<!doctype html>
<html>
<head>
    <title>Consulta New Relic - Message ID</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            background-color: #f4f6f8;
            padding: 30px;
            color: #333;
        }
        h2 {
            color: #0066cc;
        }
        form {
            margin-bottom: 20px;
        }
        input[type="text"], input[type="datetime-local"] {
            padding: 8px;
            width: 300px;
            border: 1px solid #ccc;
            border-radius: 5px;
        }
        input[type="submit"] {
            padding: 8px 15px;
            background-color: #0066cc;
            color: white;
            border: none;
            border-radius: 5px;
            cursor: pointer;
        }
        ul {
            list-style: none;
            padding: 0;
        }
        li {
            background-color: white;
            padding: 15px;
            border: 1px solid #ddd;
            margin-bottom: 10px;
            border-radius: 8px;
        }
        .not-found, .error {
            color: red;
            font-weight: bold;
        }
        .nav {
            margin-bottom: 20px;
        }
        .nav a {
            margin-right: 10px;
            color: #0066cc;
            text-decoration: none;
        }
        .nav a:hover {
            text-decoration: underline;
        }
    </style>
</head>
<body>
    <div class="nav">
        <a href="/">Consulta por Message ID</a>
        <a href="/csv-download">Consulta por Período e Company ID</a>
    </div>
    <h2>Consulta de message.id no New Relic</h2>
    <form method="post">
        <input name="message_id" type="text" placeholder="Digite o message.id" required>
        <input type="submit" value="Buscar">
    </form>

    {% if error %}
        <p class="error">{{ error }}</p>
    {% endif %}
    {% if results is not none %}
        {% if results|length == 0 %}
            <p class="not-found">message.id não encontrado.</p>
        {% else %}
            <h3>Resultados:</h3>
            <ul>
            {% for item in results %}
                <li>
                    <strong>chat.id:</strong> {{ item['chat.id'] }}<br>
                    <strong>status.code:</strong> {{ item['status.code'] }}<br>
                    <strong>status.description:</strong> {{ item['status.description'] }}<br>
                </li>
            {% endfor %}
            </ul>
        {% endif %}
    {% endif %}
</body>
</html>
"""

HTML_CSV_DOWNLOAD = """
<!doctype html>
<html>
<head>
    <title>Consulta New Relic - CSV Download</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            background-color: #f4f6f8;
            padding: 30px;
            color: #333;
        }
        h2 {
            color: #0066cc;
        }
        form {
            margin-bottom: 20px;
        }
        input[type="text"], input[type="datetime-local"] {
            padding: 8px;
            width: 300px;
            border: 1px solid #ccc;
            border-radius: 5px;
            margin-bottom: 10px;
        }
        input[type="submit"] {
            padding: 8px 15px;
            background-color: #0066cc;
            color: white;
            border: none;
            border-radius: 5px;
            cursor: pointer;
        }
        .error, .not-found {
            color: red;
            font-weight: bold;
        }
        .nav {
            margin-bottom: 20px;
        }
        .nav a {
            margin-right: 10px;
            color: #0066cc;
            text-decoration: none;
        }
        .nav a:hover {
            text-decoration: underline;
        }
    </style>
</head>
<body>
    <div class="nav">
        <a href="/">Consulta por Message ID</a>
        <a href="/csv-download">Consulta por Período e Company ID</a>
    </div>
    <h2>Consulta por Período e Company ID</h2>
    <form method="post">
        <label>Data e hora de início (máximo 24h de intervalo):</label><br>
        <input name="start_date" type="datetime-local" required><br>
        <label>Data e hora de término:</label><br>
        <input name="end_date" type="datetime-local" required><br>
        <label>Company ID:</label><br>
        <input name="company_id" type="text" placeholder="Digite o company.id" required><br>
        <input type="submit" value="Gerar CSV">
    </form>

    {% if error %}
        <p class="error">{{ error }}</p>
    {% endif %}
    {% if results is not none %}
        {% if results|length == 0 %}
            <p class="not-found">Nenhum resultado encontrado no intervalo especificado.</p>
        {% endif %}
    {% endif %}
</body>
</html>
"""


@app.route('/', methods=['GET', 'POST'])
def index():
    results = None
    error = None
    if request.method == 'POST':
        message_id = request.form.get('message_id')
        if not message_id:
            error = "message.id não pode ser vazio."
            return render_template_string(HTML_MESSAGE_ID, results=results, error=error)

        query = {
            "query": f"""{{
              actor {{
                account(id: {ACCOUNT_ID}) {{
                  nrql(query: "SELECT chat.id, status.code, status.description FROM Log WHERE (type = 'SENT_MESSAGE_STATUS' OR type = 'SENT_MESSAGE' OR type = 'SENT_ALL_MESSAGE') AND message.id IN ({message_id}) SINCE 1 month ago UNTIL now") {{
                    results
                  }}
                }}
              }}
            }}"""
        }
        try:
            response = requests.post(
                "https://api.newrelic.com/graphql",
                headers={
                    "API-Key": API_KEY,
                    "Content-Type": "application/json"
                },
                json=query
            )
            response.raise_for_status()
            try:
                data = response.json()
            except ValueError:
                error = "Erro: Resposta da API não é um JSON válido"
                return render_template_string(HTML_MESSAGE_ID, results=results, error=error)

            if not isinstance(data, dict):
                error = "Erro: Resposta da API não contém um dicionário válido"
                return render_template_string(HTML_MESSAGE_ID, results=results, error=error)

            results = data.get("data", {}).get("actor", {}).get("account", {}).get("nrql", {}).get("results", [])
        except requests.RequestException as e:
            error = f"Erro na requisição: {str(e)}"
            return render_template_string(HTML_MESSAGE_ID, results=results, error=error)

    return render_template_string(HTML_MESSAGE_ID, results=results, error=error)


@app.route('/csv-download', methods=['GET', 'POST'])
def csv_download():
    error = None
    results = None
    if request.method == 'POST':
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        company_id = request.form.get('company_id')

        # Validar company_id
        if not company_id or not company_id.strip():
            error = "Company ID não pode ser vazio."
            return render_template_string(HTML_CSV_DOWNLOAD, error=error, results=results)

        # Converter datas de entrada
        start_date, date_error = parse_brazilian_datetime(start_date_str)
        if date_error:
            error = date_error
            return render_template_string(HTML_CSV_DOWNLOAD, error=error, results=results)

        end_date, date_error = parse_brazilian_datetime(end_date_str)
        if date_error:
            error = date_error
            return render_template_string(HTML_CSV_DOWNLOAD, error=error, results=results)

        # Validar intervalo de 24 horas
        if end_date - start_date > timedelta(hours=24):
            error = "O intervalo entre as datas não pode exceder 24 horas."
            return render_template_string(HTML_CSV_DOWNLOAD, error=error, results=results)

        # Coletar resultados
        results, fetch_error = fetch_recursive(start_date, end_date, company_id)
        if fetch_error:
            error = fetch_error
            return render_template_string(HTML_CSV_DOWNLOAD, error=error, results=results)

        if results:
            # Ordenar resultados por timestamp
            results.sort(key=lambda r: int(r.get("timestamp", 0)))

            # Gerar CSV
            output = io.StringIO()
            writer = csv.writer(output, delimiter=';')
            writer.writerow(["timestamp", "chat.id", "status.description", "company.id"])
            tz_br = timezone(timedelta(hours=-3))
            for r in results:
                timestamp_ms = r.get("timestamp", "")
                chat_id = r.get("chat.id", "")
                status_desc = r.get("status.description", "")
                if timestamp_ms:
                    try:
                        timestamp_sec = int(timestamp_ms) / 1000
                        dt_utc = datetime.fromtimestamp(timestamp_sec, tz=timezone.utc)
                        dt_br = dt_utc.astimezone(tz_br)
                        timestamp_str = dt_br.strftime('%d/%m/%Y %H:%M:%S')
                    except ValueError:
                        timestamp_str = timestamp_ms
                else:
                    timestamp_str = ""
                writer.writerow([timestamp_str, chat_id, status_desc, company_id])

            # Preparar resposta para download
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype="text/csv",
                headers={"Content-Disposition": "attachment;filename=resultados.csv"}
            )

    return render_template_string(HTML_CSV_DOWNLOAD, error=error, results=results)

# --- ROTA: Relatório Huggy (Flask) ---
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import tempfile
import io
from flask import send_file

# HEADERS JIRA (mantive o mesmo Basic token que você usou antes)
HEADERS = {
    'Accept': 'application/json',
    'Content-Type': 'application/json',
    'Authorization': 'Basic dGhpYWdvLm9saXZlaXJhQGh1Z2d5LmlvOkFUQVRUM3hGZkdGMDcyNFBUR0luTDRhN0JUOGNja2Yta0R4a0hjR1ZUQnBCT3NRR1Y2ektPaGJLbmpHSlZhMEp5NE4xaFBQXzk1NmRkSF9QN01HdXdZR0drWl9HcEc0QUo4UnhFRGdyTlUyano0UkdGM1c1WGpQYU9XN0lRS2VDcHpiVjQ3SzJoM2NZTTYzeVB6bUlHLWMzSG9MSTkxZTNkcXVkMHRQbWcweThRa0pGdV9VLVBXQT00N0Q4NDVGMQ=='
}

HTML_RELATORIO = """
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Relatório Huggy</title>
  <style>
    body{font-family:Arial,Helvetica,sans-serif;background:#f7f9fb;padding:30px;color:#222}
    .card{background:#fff;padding:18px;border-radius:8px;box-shadow:0 2px 6px rgba(0,0,0,0.06);max-width:780px}
    input, label {display:block;margin-bottom:10px}
    input[type="date"], input[type="file"]{padding:8px;width:320px}
    button{padding:10px 14px;background:#0066cc;color:#fff;border:none;border-radius:6px;cursor:pointer}
    .nav{margin-bottom:12px}
    .nav a{margin-right:8px;color:#0066cc;text-decoration:none}
    .info{margin-top:12px;padding:10px;background:#eef6ff;border-radius:6px}
  </style>
</head>
<body>
  <div class="card">
    <div class="nav">
      <a href="/">Consulta Message ID</a> | <a href="/csv-download">CSV Download</a> | <strong>Relatório Huggy</strong>
    </div>
    <h2>Gerar Relatório de Atendimentos</h2>
    <form method="post" enctype="multipart/form-data">
      <label>Data inicial</label>
      <input type="date" name="start_date" required>
      <label>Data final</label>
      <input type="date" name="end_date" required>
      <label>Arquivo Data (.xlsx) - planilha "Data"</label>
      <input type="file" name="data_file" accept=".xlsx" required>
      <div style="margin-top:10px">
        <button type="submit">Gerar e Baixar Excel</button>
      </div>
    </form>

    {% if error %}
      <div class="info" style="color:#a00">{{ error }}</div>
    {% endif %}
    {% if info %}
      <div class="info">{{ info }}</div>
    {% endif %}
  </div>
</body>
</html>
"""

def get_jql_total(url):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10)
        if resp.status_code != 200:
            return 0
        data = resp.json()
        return len(data.get('issues', []))
    except Exception:
        return 0

@app.route('/relatorio-huggy', methods=['GET', 'POST'])
def relatorio_huggy():
    error = None
    info = None

    if request.method == 'POST':
        # inputs
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        uploaded = request.files.get('data_file')

        if not start_date or not end_date:
            error = "As datas inicial e final são obrigatórias."
            return render_template_string(HTML_RELATORIO, error=error, info=None)

        try:
            dt_start = datetime.strptime(start_date, "%Y-%m-%d")
            dt_end = datetime.strptime(end_date, "%Y-%m-%d")
        except Exception as e:
            error = "Formato de data inválido."
            return render_template_string(HTML_RELATORIO, error=error, info=None)

        if dt_start > dt_end:
            error = "A data inicial não pode ser maior que a final."
            return render_template_string(HTML_RELATORIO, error=error, info=None)

        if uploaded is None or uploaded.filename == "":
            error = "Envie o arquivo data_file (.xlsx)."
            return render_template_string(HTML_RELATORIO, error=error, info=None)

        # salvar arquivo temporário em memória
        try:
            df_data = pd.read_excel(uploaded, sheet_name='Data')
        except Exception as e:
            error = f"Erro ao ler o arquivo enviado: {e}"
            return render_template_string(HTML_RELATORIO, error=error, info=None)

        # lógica do relatório (mesma do script original)
        agentes_desejados = [
            'rayane.souza@huggy.io', 'antonio.carlos@huggy.io', 'henrique.ramos@huggy.io',
            'alexandre.melo@huggy.io', 'jose.sandoval@huggy.io', 'davi.nascimento@huggy.io',
            'joao.batista@huggy.io', 'thiago.oliveira@huggy.io', 'marcos.bebiano@huggy.io'
        ]

        df_filtrado = df_data[df_data['agent_login'].isin(agentes_desejados)].copy()
        total_chats = len(df_filtrado)

        df_filtrado['nota_2025'] = pd.to_numeric(df_filtrado.get('nota_2025', pd.Series()), errors='coerce')
        notas_validas = df_filtrado['nota_2025'].dropna()
        notas_validas = notas_validas[notas_validas != 0]

        media = round(notas_validas.mean(), 2) if not notas_validas.empty else 0
        mediana = round(notas_validas.median(), 2) if not notas_validas.empty else 0

        # JQLs (usando strings YYYY-MM-DD)
        start_str = dt_start.strftime("%Y-%m-%d")
        end_str = dt_end.strftime("%Y-%m-%d")
        mes_anterior = (dt_end - relativedelta(months=1)).strftime('%Y-%m-%d')

        JQL_URLS = {
            'abertos_semana': f'https://huggysupport.atlassian.net/rest/api/3/search/jql?jql=created >= "{start_str}" AND created <= "{end_str}" AND project IN (AT, HUG)&fields=key',
            'fechados_semana': f'https://huggysupport.atlassian.net/rest/api/3/search/jql?jql=statuscategorychangeddate >= "{start_str}" AND statuscategorychangeddate <= "{end_str}" AND status IN (Closed, Completed, Done, Canceled, Resolved, Declined, Reproved, "Resolvido N2", Failed, Published)&fields=key',
            'abertos_30dias': f'https://huggysupport.atlassian.net/rest/api/3/search/jql?jql=project IN (AT, HUG) AND status IN ("Analysis by Support", Escalated, "In Analysis", "In Development", "On Roadmap to Dev", Open, "Queue for Analysis", "Waiting for customer", "Work in progress", "Waiting for Information", "Tarefa Tickets", Analyzing, "In Development Partners", Opened, Pending, Reopened) AND created >= "2012-01-01" AND created <= "{mes_anterior}"&fields=key',
            'abertos_15dias': f'https://huggysupport.atlassian.net/rest/api/3/search/jql?jql=project IN (AT, HUG) AND status IN ("Analysis by Support", Escalated, "In Analysis", "In Development", "On Roadmap to Dev", Open, "Queue for Analysis", "Waiting for customer", "Work in progress", "Waiting for Information", "Tarefa Tickets", Analyzing, "In Development Partners", Opened, Pending, Reopened) AND created >= "2012-01-01" AND created <= "{(dt_end - timedelta(days=15)).strftime("%Y-%m-%d")}"&fields=key',
            'total_abertos': 'https://huggysupport.atlassian.net/rest/api/3/search/jql?jql=project IN (AT, HUG) AND status NOT IN (Closed, Done, Resolved, Reproved, "Resolvido N2", Canceled)&fields=key'
        }

        tickets_abertos_semana = get_jql_total(JQL_URLS['abertos_semana'])
        tickets_fechados_semana = get_jql_total(JQL_URLS['fechados_semana'])
        tickets_abertos_15dias = get_jql_total(JQL_URLS['abertos_15dias'])
        tickets_abertos_30dias = get_jql_total(JQL_URLS['abertos_30dias'])
        total_tickets_abertos = get_jql_total(JQL_URLS['total_abertos'])

        # checar base.xlsx na raiz do projeto
        BASE_FILE = 'base.xlsx'
        if not os.path.exists(BASE_FILE):
            error = "Arquivo base.xlsx não encontrado no servidor (coloque base.xlsx na raiz do deploy)."
            return render_template_string(HTML_RELATORIO, error=error, info=None)

        try:
            wb = load_workbook(BASE_FILE)
            ws = wb['Resultado']

            ws['B1'] = media
            ws['B2'] = mediana
            ws['B4'] = total_chats
            ws['B6'] = tickets_abertos_semana
            ws['B8'] = tickets_fechados_semana
            ws['B10'] = tickets_abertos_15dias
            ws['B11'] = tickets_abertos_30dias
            ws['B13'] = total_tickets_abertos
            ws['A15'] = f"Semana do dia {dt_start.strftime('%d/%m')} a {dt_end.strftime('%d/%m')}"

            if 'Data' in wb.sheetnames:
                wb.remove(wb['Data'])
            ws_data = wb.create_sheet('Data')
            for r in dataframe_to_rows(df_data, index=False, header=True):
                ws_data.append(r)

            # salvar em BytesIO para enviar como download
            bio = io.BytesIO()
            wb.save(bio)
            wb.close()
            bio.seek(0)

            filename = f"Atendimentos_{dt_start.strftime('%d')}_a_{dt_end.strftime('%d')}_de_{dt_end.strftime('%B')}.xlsx"
            return send_file(
                bio,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            error = f"Erro ao gerar relatório: {e}"
            return render_template_string(HTML_RELATORIO, error=error, info=None)

    # GET
    return render_template_string(HTML_RELATORIO, error=None, info=None)



if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)

